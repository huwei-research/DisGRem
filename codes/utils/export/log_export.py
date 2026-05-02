"""
log_export.py - Export logs to Excel, TXT; merge multiple runs.
Ported from MATLAB: write_log_to_excel.m, write_txt_summary.m,
                    write_sparse_log.m, merge_logs.m
"""

from __future__ import annotations
import os
import warnings
import numpy as np
from typing import List, Dict

_MONOTONE_FIELDS = {"relF", "relX", "combo", "gradNrm"}


def _running_min(arr: np.ndarray) -> np.ndarray:
    """Element-wise running minimum, NaN-aware."""
    out = arr.copy()
    best = np.inf
    for i in range(len(out)):
        if np.isfinite(out[i]):
            best = min(best, out[i])
            out[i] = best
    return out


# ─────────────────────────────────────────────────────────────
#  merge_logs
# ─────────────────────────────────────────────────────────────
def merge_logs(log_list: List[Dict], use_worst: bool = False,
               f0: float = None, f_star: float = None,
               use_median: bool = False) -> Dict:
    """
    Merge multiple per-run log dicts (one per random start) by averaging
    (or worst-case or median) across runs.
    """
    if not log_list:
        return {}

    alg_names = [k for k in log_list[0] if isinstance(log_list[0][k], dict)]
    n_run = len(log_list)
    merged = {}

    for an in alg_names:
        runs = [lg[an] for lg in log_list if an in lg]
        num_fields = ["ValueF", "gradNrm", "cons", "combo",
                      "relF", "relX", "commCost", "Mavg", "timeCost"]
        max_K = max(len(r.get("ValueF", [])) for r in runs)
        m = {}

        for fname in num_fields:
            mat = np.full((max_K, n_run), np.nan)
            for ri, r in enumerate(runs):
                if fname in r:
                    v = np.asarray(r[fname]).ravel()
                    if fname in _MONOTONE_FIELDS:
                        v = _running_min(v)
                    mat[:len(v), ri] = v
            with np.errstate(all="ignore"), warnings.catch_warnings():
                warnings.simplefilter("ignore", RuntimeWarning)
                if use_worst:
                    m[fname] = np.nanmax(mat, axis=1)
                elif use_median:
                    m[fname] = np.nanmedian(mat, axis=1)
                else:
                    m[fname] = np.nanmean(mat, axis=1)
                if fname in _MONOTONE_FIELDS:
                    m[fname] = _running_min(m[fname])
                # Store std for shaded-band plotting (only useful when n_run > 1)
                if n_run > 1:
                    m[fname + "_std"] = np.nanstd(mat, axis=1)
                    m[fname + "_q25"] = np.nanpercentile(mat, 25, axis=1)
                    m[fname + "_q75"] = np.nanpercentile(mat, 75, axis=1)
                    if fname in _MONOTONE_FIELDS:
                        m[fname + "_q25"] = _running_min(m[fname + "_q25"])
                        m[fname + "_q75"] = _running_min(m[fname + "_q75"])

        # relF fallback
        if np.all(np.isnan(np.asarray(m.get("relF", [np.nan])).ravel())):
            if f_star is not None and f0 is not None:
                denom = max(abs(f0 - f_star), 1e-12)
                m["relF"] = np.maximum((m["ValueF"] - f_star) / denom, 0)

        # Propagate __failed__ flag: mark merged log if ALL runs failed/diverged
        all_failed = all(bool(r.get("fail") or r.get("__failed__")) for r in runs)
        if all_failed:
            m["__failed__"] = True
        # Mark as "partial failure" if any (but not all) runs failed
        elif any(bool(r.get("fail") or r.get("__failed__")) for r in runs):
            m["__partial_fail__"] = True

        # xBar averaging
        if "xBar" in runs[0]:
            d = runs[0]["xBar"].shape[1] if runs[0]["xBar"].ndim == 2 else len(runs[0]["xBar"][0])
            xb_stack = np.full((max_K, d, n_run), np.nan)
            for ri, r in enumerate(runs):
                if "xBar" in r:
                    xb = np.asarray(r["xBar"])
                    T = min(len(xb), max_K)
                    xb_stack[:T, :, ri] = xb[:T]
            with np.errstate(all="ignore"):
                if use_worst:
                    m["xBar"] = np.nanmax(xb_stack, axis=2)
                else:
                    m["xBar"] = np.nanmean(xb_stack, axis=2)

        m["fail"] = any(r.get("fail", False) for r in runs)
        merged[an] = m

    return merged


# ─────────────────────────────────────────────────────────────
#  write_txt_summary
# ─────────────────────────────────────────────────────────────
def _fmt(v, fmt: str = ".2e", na: str = "---") -> str:
    """Format a numeric value; return *na* string for NaN / Inf."""
    if v is None or not np.isfinite(v):
        return na
    return f"{v:{fmt}}"


def write_txt_summary(results_dir: str, log: Dict, alg_bank: list,
                      obj_name: str, desc: str, prm: dict,
                      f0: float, f_star: float) -> None:
    """Write a human-readable TXT summary for one objective."""
    os.makedirs(results_dir, exist_ok=True)
    fpath = os.path.join(results_dir, f"summary_{obj_name}.txt")
    fpath_master = os.path.join(results_dir, "summary.txt")

    # Dynamic algorithm-name column width
    aw = max(16, max((len(a) for a, _ in alg_bank), default=16) + 1)

    lines = []
    lines.append(f"Objective: {obj_name}")
    lines.append(f"Init strategy: {desc}  (nStart={prm.get('nStart', 1)})")
    _M = prm.get("M")
    _alpha = prm.get("alpha")
    lines.append(f"Settings: N={prm.get('Nagent')}, d={prm.get('dim')}, "
                 f"M={'N/A' if _M is None else f'{_M:.2e}'}, "
                 f"alpha={'N/A' if _alpha is None else f'{_alpha:.2e}'}")
    lines.append(f"Initial f(x0): {f0:.6e}")
    lines.append(f"Reference f*:  {f_star:.6e}")
    lines.append(f"Initial gap:   {(f0 - f_star) / max(abs(f_star), 1e-12):.2e}")
    lines.append("")

    header = (f"{'Algorithm':<{aw}} | {'Time':>7} | {'Steps':>6} | "
              f"{'min(combo)':>11} | {'StdCombo':>10} | {'Converged':>9} | "
              f"{'FinalRelF':>10} | {'min(relF)':>9} | {'min(relX)':>9} | {'AvgComm(MB)':>11}")
    lines.append(header)
    lines.append("-" * len(header))

    tol = prm.get("tol", 1e-12)
    for alg_name, _ in alg_bank:
        if alg_name not in log:
            continue
        L = log[alg_name]
        combo = np.asarray(L.get("combo", [np.nan]))
        combo_std = np.asarray(L.get("combo_std", [np.nan]))
        relF = np.asarray(L.get("relF", [np.nan]))
        relX = np.asarray(L.get("relX", [np.nan]))
        tc = np.asarray(L.get("timeCost", [np.nan]))
        comm = np.asarray(L.get("commCost", [np.nan]))

        min_combo = float(np.nanmin(combo))
        last_combo_std = float(combo_std[-1]) if len(combo_std) > 0 else np.nan
        converged = "Yes" if min_combo < tol else "No"
        final_relF = float(relF[-1]) if len(relF) > 0 else np.nan
        time_val = float(tc[-1]) if len(tc) > 0 else np.nan
        min_relF = float(np.nanmin(relF))
        min_relX = float(np.nanmin(relX))
        avg_comm = float(np.nanmean(comm))

        lines.append(
            f"{alg_name:<{aw}} | {_fmt(time_val, '7.2f'):>7} | "
            f"{len(combo):>6} | {_fmt(min_combo, '11.2e'):>11} | "
            f"{_fmt(last_combo_std, '10.2e'):>10} | {converged:>9} | "
            f"{_fmt(final_relF, '10.2e'):>10} | "
            f"{_fmt(min_relF, '9.2e'):>9} | {_fmt(min_relX, '9.2e'):>9} | "
            f"{_fmt(avg_comm, '11.2f'):>11}"
        )

    content = "\n".join(lines) + "\n"
    with open(fpath, "w", encoding="utf-8") as fh:
        fh.write(content)
    print(f"[Saved] {fpath}")
    with open(fpath_master, "a", encoding="utf-8") as fh:
        fh.write(content + "\n")


# ─────────────────────────────────────────────────────────────
#  write_log_to_excel
# ─────────────────────────────────────────────────────────────
def write_log_to_excel(logs_each: List[Dict], alg_bank: list,
                       obj_name: str, prm: dict,
                       f0: float, f_star: float, filename: str) -> None:
    """
    Export sparse per-iteration summary to an Excel workbook.
    One sheet per algorithm, ~20 log-spaced iterations.
    Requires openpyxl.
    """
    try:
        import openpyxl
    except ImportError:
        print("[Warning] openpyxl not installed - skipping Excel export.")
        return

    os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)

    try:
        wb = openpyxl.load_workbook(filename)
    except Exception:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    K_max = prm.get("maxIt", 500)
    n_start = len(logs_each)
    print_idx = np.unique(np.round(
        np.logspace(0, np.log10(max(K_max, 1)), 20)).astype(int))
    print_idx = print_idx[(print_idx >= 1) & (print_idx <= K_max)]

    header = ["init_id", "iter", "ValueF", "gradNrm", "cons", "combo",
              "relF", "relX", "commCost(MB)", "timeCost(s)"]

    for alg_name, _ in alg_bank:
        sheet_name = alg_name[:31]
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(header)

        for s, log_s in enumerate(logs_each):
            if alg_name not in log_s:
                continue
            blk = log_s[alg_name]
            for k in print_idx:
                idx = int(k) - 1
                row = [s + 1, int(k)]
                for field in ["ValueF", "gradNrm", "cons", "combo",
                              "relF", "relX", "commCost", "timeCost"]:
                    v = blk.get(field, [])
                    row.append(float(v[idx]) if idx < len(v) else None)
                ws.append(row)

    wb.save(filename)
    print(f"[Saved] {filename}")
